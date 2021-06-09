import paramiko
import time
import openpyxl
hostname = "vmaistest12.oracleoutsourcing.com"
#############################################################Open Excel file and Write To Text File#########################################################
data = ""
l=[]
f = open("abhi.txt","w")
wb = openpyxl.load_workbook('a.xlsx')
sheet = wb.get_sheet_by_name('Worksheet')
#print sheet.max_row
#print sheet.max_column
for i in range(2,sheet.max_row):
    for j in [1,2,5]:
        l.append(sheet.cell(row=i+1, column=j).value)
    for item in l:
        f.write("%s\t" % item)
    f.write("\n")
    l = []
f.close()
########################################################Open Text File and Create a Dictionary##############################################################
lis = []
d = {}
f = open("abhi.txt",'r')
for line in f:
    lis = line.split()
    if d.has_key(lis[2]):
        d[lis[2]] = d[lis[2]] + "/" + lis[0]
    else:
        d[lis[2]] = lis[0]
f.close()
#####################################################Display Contents of Dictionary#########################################################################
f = open('output.txt','w')
m = d.keys()
n = d.values()
o = []
for i in n:
    temp=(i.split('/'))
    o.append(len(temp))
for i,j,k in zip(m,n,o):
    string = i + "~~" + j + ":" + "Total: " + str(k) + '\n'
    f.write(string)

f.close()

########################################################Start the paramiko SSH client######################################################################
ssh = paramiko.SSHClient()
ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
try:
    ssh.connect(hostname, username='root', password='welcome1')
    channel = ssh.invoke_shell()
    channel.settimeout(120)
    
    sftp = ssh.open_sftp()
    sftp.put('C:/Users/abhr/Desktop/mail.py', '/root/mail.py' )
    sftp.put('C:/Users/abhr/Desktop/output.txt', '/root/file.txt' )
    sftp.close()
###############################################Execute the test script to findout the error###############################################################
    channel.send('python /root/mail.py '+'\n')
    buff = ''
    while not buff.endswith('# '):
        resp = channel.recv(9999)
        buff += resp

    time.sleep(1)
    channel.send('rm -rf /root/mail.py /root/file.txt' + '\n')
    buff = ''
    while not buff.endswith('# '):
        resp = channel.recv(9999)
        buff += resp

    time.sleep(1)
                      
except paramiko.SSHException:
	print "SSH Authentication Failed"
	quit()


ssh.close()

