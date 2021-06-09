############################Open Excel file and Write To Text File###############################
import openpyxl
data = ""
l=[]
f = open("abhi.txt","w")
wb = openpyxl.load_workbook('a.xlsx')
sheet = wb.get_sheet_by_name('Worksheet')
#print sheet.max_row
#print sheet.max_column
for i in range(2,sheet.max_row):
    for j in [1,2,5]:
        l.append(sheet.cell(row=i+1, column=j).value)
    for item in l:
        f.write("%s\t" % item)
    f.write("\n")
    l = []
f.close()
###########################Open Text File and Create a Dictionary#################################
lis = []
d = {}
f = open("abhi.txt",'r')
for line in f:
    lis = line.split()
    if d.has_key(lis[2]):
        d[lis[2]] = d[lis[2]] + "/" + lis[0]
    else:
        d[lis[2]] = lis[0]
f.close()
############################Display Contents of Dictionary########################################
f = open('output.txt','w')
m = d.keys()
n = d.values()
o = []
for i in n:
    temp=(i.split('/'))
    o.append(len(temp))
for i,j,k in zip(m,n,o):
    string = i + "~~" + j + ":" + "Total: " + str(k) + '\n'
    f.write(string)

f.close()
##############################Send Mail############################################################

